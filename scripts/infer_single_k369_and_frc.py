# -*- coding: utf-8 -*-
"""
infer_single_k369_and_frc.py

Generate single-image reconstructions for K=3/6/9 from one GT image,
save K3/K6/K9/ours_norm01.tif, save y_full_norm01.tif (from K9 raw stack),
and compute GT-vs-reconstruction FRC-style curves plus a WF baseline.

IMPORTANT
---------
1) This script is intended for figure generation / single-example analysis.
2) The FRC curves produced here follow the same GT-vs-reconstruction logic as
   your current frc_gt_batch.py. They are useful for visualisation, but in the
   rebuttal and manuscript you should additionally report dataset-level paired
   FRC computed from two independent reconstructions per sample.
3) This script assumes the project modules model_recon_scunet.py exist in the
   same project environment and that the checkpoints are compatible.
"""

from __future__ import annotations

import math
import os
import random
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import tifffile as tiff

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

import torch
import torch.nn as nn

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from model_recon_scunet import SCUNetRecon


class CFG:
    DEVICE = "cuda" if torch.cuda.is_available() else "cpu"
    CHANNELS_LAST = True
    SEED = 2024

    # ---------- input / output ----------
    GT_PATH = r"C:\Users\1\OneDrive\Desktop\ceshiji\microtubules_Cell_055_SIM_gt.tif"
    OUT_DIR = r"E:\code\RL-SIM1\single_k369_out\microtubules_Cell_055"

    # ---------- checkpoints ----------
    CKPT_K3 = r"E:\code\RL-SIM1\work_all\APCRL_Zoff_K3_angles_minmax\best_ssim.pth"
    CKPT_K6 = r"E:\code\RL-SIM1\work_all\APCRL_Zoff_K6_angles_minmax\best_ssim.pth"
    CKPT_K9 = r"E:\code\RL-SIM1\work_all\APCRL_Zoff_K9_angles_minmax\best_ssim.pth"

    # ---------- forward model ----------
    BASE_CYCLES = 24.0
    OTF_SIGMA = 0.18
    PATCH_SIZE = 256  # set to None to use full image

    # ---------- FRC / WF config ----------
    PIXEL_SIZE_NM = 31.3
    PIXEL_UNIT = "nm"
    DETECTION_NA = 1.3
    WAVELENGTH_NM = 488.0
    ROI = None  # e.g. (x, y, w, h)

    APOD_PX = 20
    REMOVE_MEAN = True
    MIN_RING_SAMPLES = 64
    SMOOTH_WIN = 9
    DPI = 300


# =============================================================================
# Basic utilities
# =============================================================================
def set_seed(seed: int) -> None:
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(seed)


def imread(path: str) -> np.ndarray:
    arr = tiff.imread(path)
    return np.asarray(arr)


def norm01(arr: np.ndarray) -> np.ndarray:
    x = arr.astype(np.float32)
    if x.ndim == 3:
        x = x.mean(axis=2)
    vmax = float(x.max()) if x.size else 1.0
    vmax = max(vmax, 1e-6)
    return np.clip(x / vmax, 0.0, 1.0)


def to_tensor_1chw(x01: np.ndarray) -> torch.Tensor:
    return torch.from_numpy(x01[None, ...]).float()


def center_crop(x: torch.Tensor, patch: Optional[int]) -> torch.Tensor:
    if patch is None:
        return x
    C, H, W = x.shape
    if H <= patch or W <= patch:
        return x
    top = (H - patch) // 2
    left = (W - patch) // 2
    return x[:, top:top + patch, left:left + patch]


def save_tif(path: Path, arr: np.ndarray) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tiff.imwrite(str(path), arr.astype(np.float32), photometric="minisblack")


def read_tif_2d(path: str) -> np.ndarray:
    arr = tiff.imread(path)
    if arr.ndim == 2:
        return arr.astype(np.float32, copy=False)
    if arr.ndim == 3:
        return arr[0].astype(np.float32, copy=False)
    raise ValueError(f"Expected 2D/3D tif, got shape={arr.shape} from {path}")


def apply_roi(x: np.ndarray, roi: Optional[Tuple[int, int, int, int]]) -> np.ndarray:
    if roi is None:
        return x
    x0, y0, w, h = roi
    if w <= 0 or h <= 0:
        raise ValueError(f"Invalid ROI: {roi}")
    if (x0 < 0) or (y0 < 0) or (x0 + w > x.shape[1]) or (y0 + h > x.shape[0]):
        raise ValueError(f"ROI {roi} out of bounds for shape {x.shape}")
    return x[y0:y0 + h, x0:x0 + w]


def center_crop_to_size(x: np.ndarray, h: int, w: int) -> np.ndarray:
    H, W = x.shape
    if h > H or w > W:
        raise ValueError(f"Requested crop {(h, w)} exceeds image shape {(H, W)}")
    y0 = (H - h) // 2
    x0 = (W - w) // 2
    return x[y0:y0 + h, x0:x0 + w]


def autosize_worksheet(ws) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 70)


# =============================================================================
# SCUNet loading (copied from eval_fig1_k369_angles_minmax.py)
# =============================================================================
def inflate_in_channels(w0: torch.Tensor, in_ch_target: int) -> torch.Tensor:
    out_c, C0, kh, kw = w0.shape
    if in_ch_target == C0:
        return w0.clone()
    if in_ch_target > C0:
        rep = math.ceil(in_ch_target / C0)
        w_cat = w0.repeat(1, rep, 1, 1)[:, :in_ch_target]
        scale = math.sqrt(C0 / float(in_ch_target))
        return w_cat * scale
    idxs = torch.linspace(0, C0, steps=in_ch_target + 1, device=w0.device).round().long().tolist()
    chunks = []
    for i in range(in_ch_target):
        l, r = idxs[i], idxs[i + 1]
        if r <= l:
            r = min(l + 1, C0)
        chunks.append(w0[:, l:r].mean(dim=1, keepdim=True))
    new_w = torch.cat(chunks, dim=1)
    scale = math.sqrt(C0 / float(in_ch_target))
    return new_w * scale


def replace_conv(parent: nn.Module, name: str, conv: nn.Conv2d, new_in_ch: int):
    new_conv = nn.Conv2d(
        new_in_ch,
        conv.out_channels,
        kernel_size=conv.kernel_size,
        stride=conv.stride,
        padding=conv.padding,
        dilation=conv.dilation,
        groups=1,
        bias=(conv.bias is not None),
        padding_mode=conv.padding_mode,
    ).to(device=conv.weight.device, dtype=conv.weight.dtype)
    with torch.no_grad():
        new_conv.weight.copy_(inflate_in_channels(conv.weight.data, new_in_ch).to(conv.weight.dtype))
        if conv.bias is not None:
            new_conv.bias.copy_(conv.bias.data)
    parent._modules[name] = new_conv


def adapt_input_convs(net: nn.Module, target_in_ch: int) -> nn.Module:
    if hasattr(net, "head"):
        h = getattr(net, "head")
        if isinstance(h, nn.Sequential) and len(h) > 0 and isinstance(h[0], nn.Conv2d):
            if h[0].in_channels != target_in_ch:
                replace_conv(h, "0", h[0], target_in_ch)
        elif isinstance(h, nn.Conv2d):
            if h.in_channels != target_in_ch:
                replace_conv(net, "head", h, target_in_ch)

    def dfs(parent: nn.Module):
        for name, child in parent.named_children():
            if isinstance(child, nn.Conv2d) and child.in_channels == 6 and target_in_ch != 6:
                replace_conv(parent, name, child, target_in_ch)
            else:
                dfs(child)

    dfs(net)
    return net


def load_scunet(ckpt: str, in_ch: int, device: torch.device) -> nn.Module:
    ckpt_obj = torch.load(ckpt, map_location=device)
    state = ckpt_obj["model"] if isinstance(ckpt_obj, dict) and "model" in ckpt_obj else ckpt_obj
    net = SCUNetRecon(base_ch=48).to(device)
    if in_ch != 6:
        net = adapt_input_convs(net, in_ch)
    if CFG.CHANNELS_LAST:
        net = net.to(memory_format=torch.channels_last)
    net.load_state_dict(state, strict=False)
    net.eval()
    return net


# =============================================================================
# Forward model (copied from eval_fig1_k369_angles_minmax.py)
# =============================================================================
class SIMForwardAngles(nn.Module):
    def __init__(self, angles_deg: List[float], base_cycles: float, otf_sigma: float):
        super().__init__()
        self.angles_deg = [float(a) for a in angles_deg]
        self.base_cycles = float(base_cycles)
        self.otf_sigma = float(otf_sigma)

    def _build_otf(self, H: int, W: int, device: torch.device):
        ys = torch.linspace(-(H - 1) / 2, (H - 1) / 2, steps=H, device=device, dtype=torch.float32)
        xs = torch.linspace(-(W - 1) / 2, (W - 1) / 2, steps=W, device=device, dtype=torch.float32)
        yy, xx = torch.meshgrid(ys, xs, indexing="ij")
        rr2 = (xx ** 2 + yy ** 2)
        rr2 = rr2 / (rr2.max() + 1e-6)
        A = torch.exp(-rr2 / (2.0 * (self.otf_sigma ** 2))).contiguous()
        return A

    def _apply_otf(self, img: torch.Tensor, A: torch.Tensor) -> torch.Tensor:
        X = torch.fft.fft2(img.float(), dim=(-2, -1), norm="ortho")
        X = torch.fft.fftshift(X, dim=(-2, -1))
        Y = X * A
        Y = torch.fft.ifftshift(Y, dim=(-2, -1))
        y = torch.fft.ifft2(Y, dim=(-2, -1), norm="ortho").real
        return y.to(img.dtype)

    def forward(self, gt: torch.Tensor) -> torch.Tensor:
        B, C, H, W = gt.shape
        device = gt.device
        A = self._build_otf(H, W, device)

        xs = torch.linspace(0, float(W - 1), steps=W, device=device).view(1, 1, 1, W)
        ys = torch.linspace(0, float(H - 1), steps=H, device=device).view(1, 1, H, 1)
        k_mag = 2.0 * math.pi * float(CFG.BASE_CYCLES) / float(W)

        base_ph = torch.tensor([0.0, 2 * math.pi / 3, 4 * math.pi / 3], device=device, dtype=torch.float32)

        frames = []
        for j in range(3):
            phi = base_ph[j]
            for ang in self.angles_deg:
                th = math.radians(ang)
                kx = k_mag * math.cos(th)
                ky = k_mag * math.sin(th)
                patt = 1.0 + torch.cos(kx * xs + ky * ys + phi)
                img = (gt * patt).clamp(0, 1)
                img = self._apply_otf(img, A)
                frames.append(img)
        return torch.cat(frames, dim=1)


# =============================================================================
# Single-image inference
# =============================================================================
@torch.no_grad()
def infer_one(
    net: nn.Module,
    forward: SIMForwardAngles,
    gt_path: str,
    device: torch.device,
) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    arr = imread(gt_path)
    x01 = norm01(arr)
    ten = to_tensor_1chw(x01)
    ten = center_crop(ten, CFG.PATCH_SIZE)
    gt = ten.unsqueeze(0).to(device)
    if CFG.CHANNELS_LAST:
        gt = gt.contiguous(memory_format=torch.channels_last)

    raw = forward(gt)
    if CFG.CHANNELS_LAST:
        raw = raw.contiguous(memory_format=torch.channels_last)

    rec = torch.sigmoid(net(raw)).clamp(0, 1)

    gt_np = gt[0, 0].detach().cpu().numpy().astype(np.float32)
    raw_np = raw[0].detach().cpu().numpy().astype(np.float32)   # (N,H,W)
    rec_np = rec[0, 0].detach().cpu().numpy().astype(np.float32)
    return gt_np, raw_np, rec_np


# =============================================================================
# Widefield OTF (copied / adapted from frc_gt_batch.py)
# =============================================================================
def theoretical_resolutions_nm(na: float, wavelength_nm: float) -> Tuple[float, float]:
    abbe = wavelength_nm / (2.0 * na)
    rayleigh = 0.61 * wavelength_nm / na
    return float(abbe), float(rayleigh)


def widefield_cutoff_cyc_per_pix(na: float, wavelength_nm: float, pixel_size_nm: float) -> float:
    lam_um = wavelength_nm * 1e-3
    pix_um = pixel_size_nm * 1e-3
    fc_cyc_per_um = 2.0 * na / lam_um
    fc_cyc_per_pix = fc_cyc_per_um * pix_um
    return float(fc_cyc_per_pix)


def make_incoherent_otf(h: int, w: int, fc_cyc_per_pix: float) -> np.ndarray:
    if not np.isfinite(fc_cyc_per_pix) or fc_cyc_per_pix <= 0:
        raise ValueError(f"Invalid fc_cyc_per_pix={fc_cyc_per_pix}")

    fc = min(fc_cyc_per_pix, 0.5)
    fy = (np.arange(h) - (h // 2)) / float(h)
    fx = (np.arange(w) - (w // 2)) / float(w)
    FX, FY = np.meshgrid(fx, fy)
    FR = np.sqrt(FX * FX + FY * FY)

    rho = FR / fc
    otf = np.zeros((h, w), dtype=np.float64)
    m = rho <= 1.0
    r = np.clip(rho[m], 0.0, 1.0)
    otf[m] = (2.0 / math.pi) * (np.arccos(r) - r * np.sqrt(np.maximum(0.0, 1.0 - r * r)))
    return np.clip(otf, 0.0, 1.0)


def apply_otf_filter(img: np.ndarray, otf: np.ndarray) -> np.ndarray:
    if img.shape != otf.shape:
        raise ValueError(f"Shape mismatch: img={img.shape}, otf={otf.shape}")
    x = np.nan_to_num(img.astype(np.float64, copy=False), nan=0.0, posinf=0.0, neginf=0.0)
    F = np.fft.fftshift(np.fft.fft2(x))
    Ff = F * otf
    y = np.real(np.fft.ifft2(np.fft.ifftshift(Ff)))
    y = np.maximum(y, 0.0)
    return y.astype(np.float32)


# =============================================================================
# FRC computation (copied / adapted from frc_gt_batch.py)
# =============================================================================
def cosine_apodization_window(h: int, w: int, apod_px: int) -> np.ndarray:
    if apod_px <= 0:
        return np.ones((h, w), dtype=np.float64)
    ap = int(apod_px)
    ap = min(ap, h // 2 - 1, w // 2 - 1)
    if ap <= 0:
        return np.ones((h, w), dtype=np.float64)
    t = 0.5 * (1.0 - np.cos(np.linspace(0.0, math.pi, ap, dtype=np.float64)))
    wy = np.ones(h, dtype=np.float64)
    wx = np.ones(w, dtype=np.float64)
    wy[:ap] = t
    wy[-ap:] = t[::-1]
    wx[:ap] = t
    wx[-ap:] = t[::-1]
    return wy[:, None] * wx[None, :]


def smooth_1d_nan(x: np.ndarray, win: int) -> np.ndarray:
    if win <= 1:
        return x.astype(np.float64, copy=True)
    win = int(win)
    y = np.full_like(x, np.nan, dtype=np.float64)
    n = x.size
    half = win // 2
    for i in range(n):
        j0 = max(0, i - half)
        j1 = min(n, i + half + 1)
        seg = x[j0:j1]
        seg = seg[np.isfinite(seg)]
        if seg.size > 0:
            y[i] = float(np.mean(seg))
    return y


def half_bit_threshold(Ni: np.ndarray) -> np.ndarray:
    Ni = Ni.astype(np.float64)
    T = np.full_like(Ni, np.nan, dtype=np.float64)
    m = Ni > 0
    s = np.sqrt(Ni[m])
    T[m] = (0.2071 + 1.9102 / s) / (1.2701 + 0.9102 / s)
    return T


def compute_frc(
    img_ref: np.ndarray,
    img_tst: np.ndarray,
    apod_px: int,
    remove_mean: bool,
    min_ring_samples: int,
    smooth_win: int,
) -> Dict[str, np.ndarray]:
    if img_ref.shape != img_tst.shape:
        raise ValueError(f"Shape mismatch: ref={img_ref.shape}, tst={img_tst.shape}")

    ref = np.nan_to_num(img_ref.astype(np.float64, copy=False), nan=0.0, posinf=0.0, neginf=0.0)
    tst = np.nan_to_num(img_tst.astype(np.float64, copy=False), nan=0.0, posinf=0.0, neginf=0.0)

    if remove_mean:
        ref = ref - float(np.mean(ref))
        tst = tst - float(np.mean(tst))

    H, W = ref.shape
    win = cosine_apodization_window(H, W, apod_px=apod_px)
    refw = ref * win
    tstw = tst * win

    F1 = np.fft.fftshift(np.fft.fft2(refw))
    F2 = np.fft.fftshift(np.fft.fft2(tstw))

    fy = (np.arange(H) - (H // 2)) / float(H)
    fx = (np.arange(W) - (W // 2)) / float(W)
    FX, FY = np.meshgrid(fx, fy)
    FR = np.sqrt(FX * FX + FY * FY)

    f_max = 0.5
    m = FR <= f_max
    FRv = FR[m].ravel()
    F1v = F1[m].ravel()
    F2v = F2[m].ravel()

    Nmin = min(H, W)
    n_bins = int(np.floor(f_max * Nmin))
    n_bins = max(n_bins, 8)

    bin_idx = np.floor(FRv / f_max * n_bins).astype(np.int64)
    bin_idx = np.clip(bin_idx, 0, n_bins - 1)

    Ni = np.bincount(bin_idx, minlength=n_bins).astype(np.int64)

    cross = np.real(F1v * np.conj(F2v))
    p1 = np.abs(F1v) ** 2
    p2 = np.abs(F2v) ** 2

    num = np.bincount(bin_idx, weights=cross, minlength=n_bins).astype(np.float64)
    den1 = np.bincount(bin_idx, weights=p1, minlength=n_bins).astype(np.float64)
    den2 = np.bincount(bin_idx, weights=p2, minlength=n_bins).astype(np.float64)

    eps = 1e-12
    frc = num / np.sqrt(den1 * den2 + eps)
    frc = frc.astype(np.float64)
    frc[Ni < int(min_ring_samples)] = np.nan

    f_cyc_per_pix = (np.arange(n_bins, dtype=np.float64) + 0.5) * f_max / float(n_bins)
    f_norm = f_cyc_per_pix / f_max
    thr_1over7 = np.full(n_bins, 1.0 / 7.0, dtype=np.float64)
    thr_half = half_bit_threshold(Ni.astype(np.float64))
    frc_s = smooth_1d_nan(frc, win=smooth_win)

    pixel_um = CFG.PIXEL_SIZE_NM * 1e-3
    f_cyc_per_um = f_cyc_per_pix / pixel_um

    return {
        "f_norm": f_norm,
        "f_cyc_per_pix": f_cyc_per_pix,
        "f_cyc_per_um": f_cyc_per_um,
        "Ni": Ni.astype(np.float64),
        "frc_raw": frc,
        "frc_smooth": frc_s,
        "thr_halfbit": thr_half,
        "thr_1over7": thr_1over7,
    }


def estimate_cutoff(f: np.ndarray, frc: np.ndarray, thr: np.ndarray) -> Optional[float]:
    valid = np.isfinite(f) & np.isfinite(frc) & np.isfinite(thr)
    if not np.any(valid):
        return None
    good = valid & (frc >= thr)
    idx = np.where(good)[0]
    if idx.size == 0:
        return None
    i = int(idx[-1])
    if i >= f.size - 1:
        return float(f[i])
    g0 = float(frc[i] - thr[i])
    g1 = float(frc[i + 1] - thr[i + 1])
    if (not np.isfinite(g0)) or (not np.isfinite(g1)):
        return float(f[i])
    if g1 >= 0 or g0 == g1:
        return float(f[i])
    t = g0 / (g0 - g1)
    t = min(max(t, 0.0), 1.0)
    return float(f[i] + t * (f[i + 1] - f[i]))


def freq_to_resolution(f_cut: Optional[float]) -> Tuple[Optional[float], Optional[float]]:
    if f_cut is None or (not np.isfinite(f_cut)) or f_cut <= 0:
        return None, None
    res_px = 1.0 / float(f_cut)
    res_nm = res_px * float(CFG.PIXEL_SIZE_NM)
    return res_px, res_nm


def save_csv(path: Path, header: List[str], cols: List[np.ndarray]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    M = np.stack(cols, axis=1)
    np.savetxt(str(path), M, delimiter=",", header=",".join(header), comments="", fmt="%.10g")


# =============================================================================
# Main
# =============================================================================
def main() -> None:
    set_seed(CFG.SEED)
    out_dir = Path(CFG.OUT_DIR)
    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "K3").mkdir(parents=True, exist_ok=True)
    (out_dir / "K6").mkdir(parents=True, exist_ok=True)
    (out_dir / "K9").mkdir(parents=True, exist_ok=True)

    device = torch.device(CFG.DEVICE)
    print(f"[device] {device}")

    # ---------- load models ----------
    net3 = load_scunet(CFG.CKPT_K3, 3, device)
    net6 = load_scunet(CFG.CKPT_K6, 6, device)
    net9 = load_scunet(CFG.CKPT_K9, 9, device)

    f3 = SIMForwardAngles([0.0], CFG.BASE_CYCLES, CFG.OTF_SIGMA).to(device).eval()
    f6 = SIMForwardAngles([0.0, 60.0], CFG.BASE_CYCLES, CFG.OTF_SIGMA).to(device).eval()
    f9 = SIMForwardAngles([0.0, 60.0, 120.0], CFG.BASE_CYCLES, CFG.OTF_SIGMA).to(device).eval()

    # ---------- infer ----------
    gt3, raw3, rec3 = infer_one(net3, f3, CFG.GT_PATH, device)
    gt6, raw6, rec6 = infer_one(net6, f6, CFG.GT_PATH, device)
    gt9, raw9, rec9 = infer_one(net9, f9, CFG.GT_PATH, device)

    # Use K9 GT as the canonical GT (they should be identical after normalisation + crop)
    gt = gt9
    save_tif(out_dir / "gt_norm_for_model_norm01.tif", gt)
    save_tif(out_dir / "K3" / "ours_norm01.tif", rec3)
    save_tif(out_dir / "K6" / "ours_norm01.tif", rec6)
    save_tif(out_dir / "K9" / "ours_norm01.tif", rec9)

    # Save raw stacks (K,H,W)
    save_tif(out_dir / "K3" / "raw_k3_norm01.tif", raw3)
    save_tif(out_dir / "K6" / "raw_k6_norm01.tif", raw6)
    save_tif(out_dir / "y_full_norm01.tif", raw9)   # reuse K9 stack as the full stack for WF baseline

    print(f"[save] {out_dir / 'K3' / 'ours_norm01.tif'}")
    print(f"[save] {out_dir / 'K6' / 'ours_norm01.tif'}")
    print(f"[save] {out_dir / 'K9' / 'ours_norm01.tif'}")
    print(f"[save] {out_dir / 'y_full_norm01.tif'}")

    # ---------- build WF baseline ----------
    gt_use = apply_roi(gt, CFG.ROI)
    y_full = raw9.astype(np.float32, copy=False)
    if CFG.ROI is not None:
        y_full = np.stack([apply_roi(y_full[k], CFG.ROI) for k in range(y_full.shape[0])], axis=0)

    wf9_mean = np.mean(y_full, axis=0).astype(np.float32)
    fc_cyc_per_pix = widefield_cutoff_cyc_per_pix(CFG.DETECTION_NA, CFG.WAVELENGTH_NM, CFG.PIXEL_SIZE_NM)
    abbe_nm, rayleigh_nm = theoretical_resolutions_nm(CFG.DETECTION_NA, CFG.WAVELENGTH_NM)
    fc_norm_nyq1 = (fc_cyc_per_pix / 0.5)

    print("============================================================")
    print(f"[Widefield OTF] NA={CFG.DETECTION_NA} | lambda={CFG.WAVELENGTH_NM} nm | pixel={CFG.PIXEL_SIZE_NM} nm")
    print(f"  Theoretical Abbe     ~ {abbe_nm:.2f} nm (lambda/(2NA))")
    print(f"  Theoretical Rayleigh ~ {rayleigh_nm:.2f} nm (0.61lambda/NA)")
    print(f"  OTF cutoff fc        = {fc_cyc_per_pix:.6f} cyc/px (Nyq=0.5 cyc/px)")
    print(f"  cutoff (Nyq=1 norm)  = {fc_norm_nyq1:.4f}")
    print("============================================================")

    otf = make_incoherent_otf(wf9_mean.shape[0], wf9_mean.shape[1], fc_cyc_per_pix)
    wf9_wfotf = apply_otf_filter(wf9_mean, otf)

    frc_out_dir = out_dir / "frc_out_truewf_otf"
    frc_out_dir.mkdir(parents=True, exist_ok=True)
    save_tif(frc_out_dir / "WF9_mean_norm01.tif", wf9_mean)
    save_tif(frc_out_dir / "WF9_wfOTF_norm01.tif", wf9_wfotf)

    imgs: Dict[str, np.ndarray] = {
        "WF9_wfOTF": wf9_wfotf,
        "K3": apply_roi(rec3, CFG.ROI),
        "K6": apply_roi(rec6, CFG.ROI),
        "K9": apply_roi(rec9, CFG.ROI),
    }
    meta: Dict[str, str] = {
        "WF9_wfOTF": f"mean(y_full) + incoherent OTF (NA={CFG.DETECTION_NA}, lambda={CFG.WAVELENGTH_NM}nm)",
        "K3": str(out_dir / "K3" / "ours_norm01.tif"),
        "K6": str(out_dir / "K6" / "ours_norm01.tif"),
        "K9": str(out_dir / "K9" / "ours_norm01.tif"),
    }

    H_common = gt_use.shape[0]
    W_common = gt_use.shape[1]
    for _, im in imgs.items():
        H_common = min(H_common, im.shape[0])
        W_common = min(W_common, im.shape[1])

    gt_use = center_crop_to_size(gt_use, H_common, W_common)
    for label in list(imgs.keys()):
        imgs[label] = center_crop_to_size(imgs[label], H_common, W_common)

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.append([
        "label",
        "source",
        "cutoff_halfbit (cycles/pixel)",
        "cutoff_halfbit (Nyq=1)",
        "res_halfbit (pixel)",
        f"res_halfbit ({CFG.PIXEL_UNIT})",
        "cutoff_1over7 (cycles/pixel)",
        "cutoff_1over7 (Nyq=1)",
        "res_1over7 (pixel)",
        f"res_1over7 ({CFG.PIXEL_UNIT})",
        "H_used",
        "W_used",
        "pixel_size_nm",
        "NA",
        "wavelength_nm",
        "otf_cutoff_cyc_per_pix",
        "note",
    ])

    overlay_curves: Dict[str, np.ndarray] = {}
    f_norm_ref = None
    thr_half_ref = None

    for label in ["WF9_wfOTF", "K3", "K6", "K9"]:
        out = compute_frc(
            img_ref=gt_use,
            img_tst=imgs[label],
            apod_px=int(CFG.APOD_PX),
            remove_mean=bool(CFG.REMOVE_MEAN),
            min_ring_samples=int(CFG.MIN_RING_SAMPLES),
            smooth_win=int(CFG.SMOOTH_WIN),
        )

        fn = out["f_norm"]
        f = out["f_cyc_per_pix"]
        f_um = out["f_cyc_per_um"]
        frc_raw = out["frc_raw"]
        frc_s = out["frc_smooth"]
        thrh = out["thr_halfbit"]
        thr17 = out["thr_1over7"]
        Ni = out["Ni"]

        if f_norm_ref is None:
            f_norm_ref = fn
            thr_half_ref = thrh

        fcut_h = estimate_cutoff(f, frc_s, thrh)
        fcut_17 = estimate_cutoff(f, frc_s, thr17)
        res_h_px, res_h_nm = freq_to_resolution(fcut_h)
        res_17_px, res_17_nm = freq_to_resolution(fcut_17)

        csv_path = frc_out_dir / f"frc_{label}.csv"
        header = [
            "f_norm(Nyq=1)",
            "f_cyc_per_pix",
            "f_cyc_per_um",
            "FRC_raw",
            "FRC_smooth",
            "thr_halfbit",
            "thr_1over7",
            "Ni",
        ]
        cols = [fn, f, f_um, frc_raw, frc_s, thrh, thr17, Ni]
        save_csv(csv_path, header, cols)

        ws = wb.create_sheet(title=label)
        ws.append(header)
        for i in range(fn.size):
            ws.append([
                float(fn[i]),
                float(f[i]),
                float(f_um[i]),
                float(frc_raw[i]) if np.isfinite(frc_raw[i]) else None,
                float(frc_s[i]) if np.isfinite(frc_s[i]) else None,
                float(thrh[i]) if np.isfinite(thrh[i]) else None,
                float(thr17[i]) if np.isfinite(thr17[i]) else None,
                float(Ni[i]) if np.isfinite(Ni[i]) else None,
            ])
        autosize_worksheet(ws)

        plt.figure(figsize=(7.2, 5.4))
        plt.plot(fn, frc_raw, linewidth=1.2, alpha=0.45, label=f"{label} raw")
        plt.plot(fn, frc_s, linewidth=2.2, label=f"{label} smooth")
        plt.plot(fn, thrh, linestyle="--", linewidth=1.4, label="half-bit thr")
        plt.axhline(1.0 / 7.0, linestyle=":", linewidth=1.4, label="1/7 thr")
        plt.title(f"FRC vs GT | {label}")
        plt.xlabel("Normalized spatial frequency (Nyquist=1)")
        plt.ylabel("FRC")
        plt.xlim(0.0, 1.0)
        plt.ylim(-0.05, 1.05)
        plt.grid(True, linewidth=0.4, alpha=0.5)
        plt.legend(loc="best", fontsize=8)
        plt.tight_layout()
        png_path = frc_out_dir / f"frc_{label}.png"
        plt.savefig(str(png_path), dpi=int(CFG.DPI))
        plt.close()

        note = ""
        if fcut_h is None or fcut_17 is None:
            note = "No cutoff found for one criterion."

        ws_sum.append([
            label,
            meta.get(label, ""),
            float(fcut_h) if fcut_h is not None else None,
            float(fcut_h / 0.5) if fcut_h is not None else None,
            float(res_h_px) if res_h_px is not None else None,
            float(res_h_nm) if res_h_nm is not None else None,
            float(fcut_17) if fcut_17 is not None else None,
            float(fcut_17 / 0.5) if fcut_17 is not None else None,
            float(res_17_px) if res_17_px is not None else None,
            float(res_17_nm) if res_17_nm is not None else None,
            int(H_common),
            int(W_common),
            float(CFG.PIXEL_SIZE_NM),
            float(CFG.DETECTION_NA),
            float(CFG.WAVELENGTH_NM),
            float(fc_cyc_per_pix),
            note,
        ])

        overlay_curves[label] = frc_s

        print(f"[{label}]")
        print(f"  source: {meta.get(label, '')}")
        print(f"  half-bit: cutoff f={fcut_h} cyc/px -> res={res_h_px} px, {res_h_nm} {CFG.PIXEL_UNIT}")
        print(f"     1/7 : cutoff f={fcut_17} cyc/px -> res={res_17_px} px, {res_17_nm} {CFG.PIXEL_UNIT}")
        print(f"  saved: {png_path} ; {csv_path}")

    autosize_worksheet(ws_sum)
    xlsx_path = frc_out_dir / "frc_results_truewf_otf.xlsx"
    wb.save(str(xlsx_path))

    if f_norm_ref is not None and thr_half_ref is not None:
        plt.figure(figsize=(7.4, 5.5))
        for label in ["WF9_wfOTF", "K3", "K6", "K9"]:
            plt.plot(f_norm_ref, overlay_curves[label], linewidth=2.3, label=f"{label} (smooth)")
        plt.plot(f_norm_ref, thr_half_ref, linestyle="--", linewidth=1.4, label="half-bit thr")
        plt.axhline(1.0 / 7.0, linestyle=":", linewidth=1.4, label="1/7 thr")
        plt.title("FRC vs GT (overlay) | WF9(widefield OTF) + K3/K6/K9")
        plt.xlabel("Normalized spatial frequency (Nyquist=1)")
        plt.ylabel("FRC (smoothed)")
        plt.xlim(0.0, 1.0)
        plt.ylim(-0.05, 1.05)
        plt.grid(True, linewidth=0.4, alpha=0.5)
        plt.legend(loc="best", fontsize=8)
        plt.tight_layout()
        overlay_png = frc_out_dir / "frc_overlay_truewf_otf.png"
        plt.savefig(str(overlay_png), dpi=int(CFG.DPI))
        plt.close()
        print(f"Overlay saved: {overlay_png}")

    print("\n==================== DONE ====================")
    print(f"GT       : {out_dir / 'gt_norm_for_model_norm01.tif'}")
    print(f"y_full   : {out_dir / 'y_full_norm01.tif'}")
    print(f"K3 recon : {out_dir / 'K3' / 'ours_norm01.tif'}")
    print(f"K6 recon : {out_dir / 'K6' / 'ours_norm01.tif'}")
    print(f"K9 recon : {out_dir / 'K9' / 'ours_norm01.tif'}")
    print(f"FRC OUT  : {frc_out_dir}")
    print(f"XLSX     : {xlsx_path}")


if __name__ == "__main__":
    main()
