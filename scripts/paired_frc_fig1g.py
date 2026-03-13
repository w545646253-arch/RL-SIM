# -*- coding: utf-8 -*-
"""
paired_frc_fig1g.py

Compute dataset-level paired FRC for Fig. 1(g) using two statistically independent
raw-stack realizations per sample and per method.

Outputs
-------
1) paired_frc_results.xlsx
2) fig1g_frc_resolution_halfbit_nm.png
3) supp_fig_s1_representative_paired_frc.png
4) paired_frc_per_sample.csv
5) paired_frc_source_curves.csv (representative sample)

Expected repository layout
--------------------------
repo_root/
  model_code/
    model_recon_scunet.py
  scripts/
    paired_frc_fig1g.py   <- place this file here
  minimal_example_data/
    *.tif
  checkpoints/
    APCRL_Zoff_K3_best.pth
    APCRL_Zoff_K6_best.pth
    APCRL_Zoff_K9_best.pth
"""
from __future__ import annotations

import math
import os
import random
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import tifffile as tiff

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

import torch
import torch.nn as nn
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# -----------------------------------------------------------------------------
# Path setup
# -----------------------------------------------------------------------------
THIS = Path(__file__).resolve()
REPO = THIS.parent.parent
MODEL_CODE = REPO / "model_code"
if str(MODEL_CODE) not in sys.path:
    sys.path.insert(0, str(MODEL_CODE))

from model_recon_scunet import SCUNetRecon  # noqa: E402


class CFG:
    DEVICE = "cuda" if torch.cuda.is_available() else "cpu"
    CHANNELS_LAST = True
    SEED = 2025

    # ----------------- paths -----------------
    GT_DIR = str(REPO / "minimal_example_data")
    OUT_DIR = str(REPO / "results_example" / "fig1g_paired_frc")

    CKPT_K3 = str(REPO / "checkpoints" / "APCRL_Zoff_K3_best.pth")
    CKPT_K6 = str(REPO / "checkpoints" / "APCRL_Zoff_K6_best.pth")
    CKPT_K9 = str(REPO / "checkpoints" / "APCRL_Zoff_K9_best.pth")

    # ----------------- forward model -----------------
    BASE_CYCLES = 24.0
    OTF_SIGMA = 0.18
    PATCH_SIZE = 256   # None -> use full image

    # ----------------- noise model for paired FRC -----------------
    # Choose values to match the moderate-count regime used in your simulated evaluation.
    PHOTONS_PEAK = 1000.0
    READ_STD_NORM = 0.002

    # ----------------- FRC config -----------------
    PIXEL_SIZE_NM = 31.3
    APOD_PX = 20
    REMOVE_MEAN = True
    MIN_RING_SAMPLES = 64
    SMOOTH_WIN = 9
    DPI = 300

    # Representative sample used for Supplementary Fig. S1
    REP_SAMPLE_BASENAME = "microtubules_Cell_055_SIM_gt"


def set_seed(seed: int) -> None:
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(seed)


def list_tifs(folder: str) -> List[str]:
    p = Path(folder)
    exts = {".tif", ".tiff"}
    return sorted([str(x) for x in p.iterdir() if x.suffix.lower() in exts])


def imread(path: str) -> np.ndarray:
    arr = tiff.imread(path)
    return np.asarray(arr)


def norm01(arr: np.ndarray) -> np.ndarray:
    x = arr.astype(np.float32)
    if x.ndim == 3:
        x = x.mean(axis=2)
    vmax = max(float(x.max()), 1e-6)
    return np.clip(x / vmax, 0.0, 1.0)


def to_tensor_1chw(x01: np.ndarray) -> torch.Tensor:
    return torch.from_numpy(x01[None, ...]).float()


def center_crop(x: torch.Tensor, patch: Optional[int]) -> torch.Tensor:
    if patch is None:
        return x
    C, H, W = x.shape
    if H <= patch or W <= patch:
        return x
    top = (H - patch) // 2
    left = (W - patch) // 2
    return x[:, top:top + patch, left:left + patch]


def inflate_in_channels(w0: torch.Tensor, in_ch_target: int) -> torch.Tensor:
    out_c, C0, kh, kw = w0.shape
    if in_ch_target == C0:
        return w0.clone()
    if in_ch_target > C0:
        rep = math.ceil(in_ch_target / C0)
        w_cat = w0.repeat(1, rep, 1, 1)[:, :in_ch_target]
        scale = math.sqrt(C0 / float(in_ch_target))
        return w_cat * scale
    idxs = torch.linspace(0, C0, steps=in_ch_target + 1, device=w0.device).round().long().tolist()
    chunks = []
    for i in range(in_ch_target):
        l, r = idxs[i], idxs[i + 1]
        if r <= l:
            r = min(l + 1, C0)
        chunks.append(w0[:, l:r].mean(dim=1, keepdim=True))
    new_w = torch.cat(chunks, dim=1)
    scale = math.sqrt(C0 / float(in_ch_target))
    return new_w * scale


def replace_conv(parent: nn.Module, name: str, conv: nn.Conv2d, new_in_ch: int):
    new_conv = nn.Conv2d(
        new_in_ch,
        conv.out_channels,
        kernel_size=conv.kernel_size,
        stride=conv.stride,
        padding=conv.padding,
        dilation=conv.dilation,
        groups=1,
        bias=(conv.bias is not None),
        padding_mode=conv.padding_mode,
    ).to(device=conv.weight.device, dtype=conv.weight.dtype)
    with torch.no_grad():
        new_conv.weight.copy_(inflate_in_channels(conv.weight.data, new_in_ch).to(conv.weight.dtype))
        if conv.bias is not None:
            new_conv.bias.copy_(conv.bias.data)
    parent._modules[name] = new_conv


def adapt_input_convs(net: nn.Module, target_in_ch: int) -> nn.Module:
    if hasattr(net, "head"):
        h = getattr(net, "head")
        if isinstance(h, nn.Sequential) and len(h) > 0 and isinstance(h[0], nn.Conv2d):
            if h[0].in_channels != target_in_ch:
                replace_conv(h, "0", h[0], target_in_ch)
        elif isinstance(h, nn.Conv2d):
            if h.in_channels != target_in_ch:
                replace_conv(net, "head", h, target_in_ch)

    def dfs(parent: nn.Module):
        for name, child in parent.named_children():
            if isinstance(child, nn.Conv2d) and child.in_channels == 6 and target_in_ch != 6:
                replace_conv(parent, name, child, target_in_ch)
            else:
                dfs(child)
    dfs(net)
    return net


def load_scunet(ckpt: str, in_ch: int, device: torch.device) -> nn.Module:
    ckpt_obj = torch.load(ckpt, map_location=device)
    state = ckpt_obj["model"] if isinstance(ckpt_obj, dict) and "model" in ckpt_obj else ckpt_obj
    net = SCUNetRecon(base_ch=48).to(device)
    if in_ch != 6:
        net = adapt_input_convs(net, in_ch)
    if CFG.CHANNELS_LAST:
        net = net.to(memory_format=torch.channels_last)
    net.load_state_dict(state, strict=False)
    net.eval()
    return net


class SIMForwardAngles(nn.Module):
    def __init__(self, angles_deg: List[float], base_cycles: float, otf_sigma: float):
        super().__init__()
        self.angles_deg = [float(a) for a in angles_deg]
        self.base_cycles = float(base_cycles)
        self.otf_sigma = float(otf_sigma)

    def _build_otf(self, H: int, W: int, device: torch.device):
        ys = torch.linspace(-(H - 1) / 2, (H - 1) / 2, steps=H, device=device, dtype=torch.float32)
        xs = torch.linspace(-(W - 1) / 2, (W - 1) / 2, steps=W, device=device, dtype=torch.float32)
        yy, xx = torch.meshgrid(ys, xs, indexing="ij")
        rr2 = (xx ** 2 + yy ** 2)
        rr2 = rr2 / (rr2.max() + 1e-6)
        A = torch.exp(-rr2 / (2.0 * (self.otf_sigma ** 2))).contiguous()
        return A

    def _apply_otf(self, img: torch.Tensor, A: torch.Tensor) -> torch.Tensor:
        X = torch.fft.fft2(img.float(), dim=(-2, -1), norm="ortho")
        X = torch.fft.fftshift(X, dim=(-2, -1))
        Y = X * A
        Y = torch.fft.ifftshift(Y, dim=(-2, -1))
        y = torch.fft.ifft2(Y, dim=(-2, -1), norm="ortho").real
        return y.to(img.dtype)

    def forward(self, gt: torch.Tensor) -> torch.Tensor:
        B, C, H, W = gt.shape
        device = gt.device
        A = self._build_otf(H, W, device)

        xs = torch.linspace(0, float(W - 1), steps=W, device=device).view(1, 1, 1, W)
        ys = torch.linspace(0, float(H - 1), steps=H, device=device).view(1, 1, H, 1)
        k_mag = 2.0 * math.pi * float(CFG.BASE_CYCLES) / float(W)
        base_ph = torch.tensor([0.0, 2 * math.pi / 3, 4 * math.pi / 3], device=device, dtype=torch.float32)

        frames = []
        for j in range(3):
            phi = base_ph[j]
            for ang in self.angles_deg:
                th = math.radians(ang)
                kx = k_mag * math.cos(th)
                ky = k_mag * math.sin(th)
                patt = 1.0 + torch.cos(kx * xs + ky * ys + phi)
                img = (gt * patt).clamp(0, 1)
                img = self._apply_otf(img, A)
                frames.append(img)
        return torch.cat(frames, dim=1)


@torch.no_grad()
def simulate_raw_clean(gt_path: str, forward: nn.Module, device: torch.device) -> np.ndarray:
    arr = imread(gt_path)
    x01 = norm01(arr)
    ten = to_tensor_1chw(x01)
    ten = center_crop(ten, CFG.PATCH_SIZE)
    gt = ten.unsqueeze(0).to(device)
    if CFG.CHANNELS_LAST:
        gt = gt.contiguous(memory_format=torch.channels_last)
    raw = forward(gt)
    raw_np = raw[0].detach().cpu().numpy().astype(np.float32)
    return raw_np


def add_poisson_gaussian_noise(raw_clean: np.ndarray, photons_peak: float, read_std_norm: float, seed: int) -> np.ndarray:
    rng = np.random.default_rng(seed)
    x = np.clip(raw_clean.astype(np.float64), 0.0, 1.0)
    lam = photons_peak * x
    shot = rng.poisson(lam).astype(np.float64) / float(photons_peak)
    noisy = shot + rng.normal(0.0, read_std_norm, size=x.shape)
    return np.clip(noisy, 0.0, 1.0).astype(np.float32)


@torch.no_grad()
def reconstruct_stack(net: nn.Module, raw_np: np.ndarray, device: torch.device) -> np.ndarray:
    raw = torch.from_numpy(raw_np[None]).to(device)
    if CFG.CHANNELS_LAST:
        raw = raw.contiguous(memory_format=torch.channels_last)
    rec = torch.sigmoid(net(raw)).clamp(0.0, 1.0)
    return rec[0, 0].detach().cpu().numpy().astype(np.float32)


def cosine_apodization_window(h: int, w: int, apod_px: int) -> np.ndarray:
    if apod_px <= 0:
        return np.ones((h, w), dtype=np.float64)
    ap = int(apod_px)
    ap = min(ap, h // 2 - 1, w // 2 - 1)
    if ap <= 0:
        return np.ones((h, w), dtype=np.float64)
    t = 0.5 * (1.0 - np.cos(np.linspace(0.0, math.pi, ap, dtype=np.float64)))
    wy = np.ones(h, dtype=np.float64)
    wx = np.ones(w, dtype=np.float64)
    wy[:ap] = t
    wy[-ap:] = t[::-1]
    wx[:ap] = t
    wx[-ap:] = t[::-1]
    return wy[:, None] * wx[None, :]


def smooth_1d_nan(x: np.ndarray, win: int) -> np.ndarray:
    if win <= 1:
        return x.astype(np.float64, copy=True)
    y = np.full_like(x, np.nan, dtype=np.float64)
    n = x.size
    half = win // 2
    for i in range(n):
        j0 = max(0, i - half)
        j1 = min(n, i + half + 1)
        seg = x[j0:j1]
        seg = seg[np.isfinite(seg)]
        if seg.size > 0:
            y[i] = float(np.mean(seg))
    return y


def half_bit_threshold(Ni: np.ndarray) -> np.ndarray:
    Ni = Ni.astype(np.float64)
    T = np.full_like(Ni, np.nan, dtype=np.float64)
    m = Ni > 0
    s = np.sqrt(Ni[m])
    T[m] = (0.2071 + 1.9102 / s) / (1.2701 + 0.9102 / s)
    return T


def compute_frc(img_a: np.ndarray, img_b: np.ndarray) -> Dict[str, np.ndarray]:
    if img_a.shape != img_b.shape:
        raise ValueError(f"Shape mismatch: {img_a.shape} vs {img_b.shape}")
    a = np.nan_to_num(img_a.astype(np.float64), nan=0.0, posinf=0.0, neginf=0.0)
    b = np.nan_to_num(img_b.astype(np.float64), nan=0.0, posinf=0.0, neginf=0.0)
    if CFG.REMOVE_MEAN:
        a -= float(np.mean(a))
        b -= float(np.mean(b))
    H, W = a.shape
    win = cosine_apodization_window(H, W, CFG.APOD_PX)
    a *= win
    b *= win
    F1 = np.fft.fftshift(np.fft.fft2(a))
    F2 = np.fft.fftshift(np.fft.fft2(b))

    fy = (np.arange(H) - (H // 2)) / float(H)
    fx = (np.arange(W) - (W // 2)) / float(W)
    FX, FY = np.meshgrid(fx, fy)
    FR = np.sqrt(FX * FX + FY * FY)
    f_max = 0.5
    m = FR <= f_max
    FRv = FR[m].ravel()
    F1v = F1[m].ravel()
    F2v = F2[m].ravel()

    Nmin = min(H, W)
    n_bins = max(int(np.floor(f_max * Nmin)), 8)
    bin_idx = np.floor(FRv / f_max * n_bins).astype(np.int64)
    bin_idx = np.clip(bin_idx, 0, n_bins - 1)
    Ni = np.bincount(bin_idx, minlength=n_bins).astype(np.int64)

    cross = np.real(F1v * np.conj(F2v))
    p1 = np.abs(F1v) ** 2
    p2 = np.abs(F2v) ** 2
    num = np.bincount(bin_idx, weights=cross, minlength=n_bins).astype(np.float64)
    den1 = np.bincount(bin_idx, weights=p1, minlength=n_bins).astype(np.float64)
    den2 = np.bincount(bin_idx, weights=p2, minlength=n_bins).astype(np.float64)
    frc = num / np.sqrt(den1 * den2 + 1e-12)
    frc[Ni < int(CFG.MIN_RING_SAMPLES)] = np.nan

    f_cyc_per_pix = (np.arange(n_bins, dtype=np.float64) + 0.5) * f_max / float(n_bins)
    f_norm = f_cyc_per_pix / 0.5
    thr_half = half_bit_threshold(Ni)
    thr_17 = np.full_like(f_cyc_per_pix, 1.0 / 7.0, dtype=np.float64)
    frc_s = smooth_1d_nan(frc, CFG.SMOOTH_WIN)
    return {
        "f_norm": f_norm,
        "f_cyc_per_pix": f_cyc_per_pix,
        "frc_raw": frc,
        "frc_smooth": frc_s,
        "thr_halfbit": thr_half,
        "thr_1over7": thr_17,
        "Ni": Ni.astype(np.float64),
    }


def estimate_cutoff(f: np.ndarray, frc: np.ndarray, thr: np.ndarray) -> Optional[float]:
    valid = np.isfinite(f) & np.isfinite(frc) & np.isfinite(thr)
    if not np.any(valid):
        return None
    good = valid & (frc >= thr)
    idx = np.where(good)[0]
    if idx.size == 0:
        return None
    i = int(idx[-1])
    if i >= f.size - 1:
        return float(f[i])
    g0 = float(frc[i] - thr[i])
    g1 = float(frc[i + 1] - thr[i + 1])
    if (not np.isfinite(g0)) or (not np.isfinite(g1)) or g1 >= 0 or g0 == g1:
        return float(f[i])
    t = g0 / (g0 - g1)
    t = min(max(t, 0.0), 1.0)
    return float(f[i] + t * (f[i + 1] - f[i]))


def cutoff_to_nm(f_cut: Optional[float]) -> Optional[float]:
    if f_cut is None or (not np.isfinite(f_cut)) or f_cut <= 0:
        return None
    return float((1.0 / f_cut) * CFG.PIXEL_SIZE_NM)


def autosize_worksheet(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 2, 50)


def draw_fig1g_boxplot(results: List[Dict[str, object]], out_png: Path) -> None:
    methods = ["WF", "RL-SIM-3", "RL-SIM-6", "RL-SIM-9"]
    keymap = {"WF": "WF_halfbit_nm", "RL-SIM-3": "K3_halfbit_nm", "RL-SIM-6": "K6_halfbit_nm", "RL-SIM-9": "K9_halfbit_nm"}
    data = [[float(r[keymap[m]]) for r in results if r.get(keymap[m]) is not None] for m in methods]
    plt.figure(figsize=(5.0, 4.2))
    bp = plt.boxplot(data, labels=methods, patch_artist=True, showfliers=False)
    for patch in bp["boxes"]:
        patch.set_alpha(0.35)
    plt.ylabel("FRC-derived resolution (nm)")
    plt.title("Held-out test set")
    plt.grid(axis="y", linewidth=0.4, alpha=0.5)
    plt.tight_layout()
    out_png.parent.mkdir(parents=True, exist_ok=True)
    plt.savefig(out_png, dpi=CFG.DPI)
    plt.close()


def draw_rep_overlay(rep_curves: Dict[str, Dict[str, np.ndarray]], rep_nm: Dict[str, float], out_png: Path) -> None:
    plt.figure(figsize=(5.2, 4.2))
    order = ["WF", "RL-SIM-3", "RL-SIM-6", "RL-SIM-9"]
    for name in order:
        d = rep_curves[name]
        label = f"{name} ({rep_nm[name]:.1f} nm)" if rep_nm.get(name) is not None else name
        plt.plot(d["f_norm"], d["frc_smooth"], linewidth=2.0, label=label)
    ref = rep_curves[order[0]]
    plt.axhline(1.0 / 7.0, linestyle=":", linewidth=1.3, label="1/7 threshold")
    plt.xlabel("Normalized spatial frequency (Nyquist = 1)")
    plt.ylabel("Fourier ring correlation")
    plt.xlim(0.0, 1.0)
    plt.ylim(-0.05, 1.05)
    plt.grid(True, linewidth=0.4, alpha=0.5)
    plt.legend(fontsize=8, loc="best")
    plt.tight_layout()
    out_png.parent.mkdir(parents=True, exist_ok=True)
    plt.savefig(out_png, dpi=CFG.DPI)
    plt.close()


def main() -> None:
    set_seed(CFG.SEED)
    out_dir = Path(CFG.OUT_DIR)
    out_dir.mkdir(parents=True, exist_ok=True)

    gt_paths = list_tifs(CFG.GT_DIR)
    if len(gt_paths) == 0:
        raise RuntimeError(f"No tif files found in {CFG.GT_DIR}")

    device = torch.device(CFG.DEVICE)
    net3 = load_scunet(CFG.CKPT_K3, 3, device)
    net6 = load_scunet(CFG.CKPT_K6, 6, device)
    net9 = load_scunet(CFG.CKPT_K9, 9, device)

    f3 = SIMForwardAngles([0.0], CFG.BASE_CYCLES, CFG.OTF_SIGMA).to(device).eval()
    f6 = SIMForwardAngles([0.0, 60.0], CFG.BASE_CYCLES, CFG.OTF_SIGMA).to(device).eval()
    f9 = SIMForwardAngles([0.0, 60.0, 120.0], CFG.BASE_CYCLES, CFG.OTF_SIGMA).to(device).eval()

    results = []
    rep_curves = {}
    rep_nm = {}

    wb = Workbook()
    ws = wb.active
    ws.title = "summary"
    ws.append([
        "sample",
        "WF_halfbit_nm", "WF_1over7_nm",
        "K3_halfbit_nm", "K3_1over7_nm",
        "K6_halfbit_nm", "K6_1over7_nm",
        "K9_halfbit_nm", "K9_1over7_nm",
    ])

    rng = np.random.default_rng(CFG.SEED)

    for idx, gt_path in enumerate(gt_paths):
        stem = Path(gt_path).stem
        # Clean forward stacks
        raw3_clean = simulate_raw_clean(gt_path, f3, device)
        raw6_clean = simulate_raw_clean(gt_path, f6, device)
        raw9_clean = simulate_raw_clean(gt_path, f9, device)

        # Two statistically independent raw stacks
        seed_base = int(rng.integers(1, 10_000_000))
        raw3_A = add_poisson_gaussian_noise(raw3_clean, CFG.PHOTONS_PEAK, CFG.READ_STD_NORM, seed_base + 11)
        raw3_B = add_poisson_gaussian_noise(raw3_clean, CFG.PHOTONS_PEAK, CFG.READ_STD_NORM, seed_base + 12)
        raw6_A = add_poisson_gaussian_noise(raw6_clean, CFG.PHOTONS_PEAK, CFG.READ_STD_NORM, seed_base + 21)
        raw6_B = add_poisson_gaussian_noise(raw6_clean, CFG.PHOTONS_PEAK, CFG.READ_STD_NORM, seed_base + 22)
        raw9_A = add_poisson_gaussian_noise(raw9_clean, CFG.PHOTONS_PEAK, CFG.READ_STD_NORM, seed_base + 31)
        raw9_B = add_poisson_gaussian_noise(raw9_clean, CFG.PHOTONS_PEAK, CFG.READ_STD_NORM, seed_base + 32)

        # Reconstructions
        rec3_A = reconstruct_stack(net3, raw3_A, device)
        rec3_B = reconstruct_stack(net3, raw3_B, device)
        rec6_A = reconstruct_stack(net6, raw6_A, device)
        rec6_B = reconstruct_stack(net6, raw6_B, device)
        rec9_A = reconstruct_stack(net9, raw9_A, device)
        rec9_B = reconstruct_stack(net9, raw9_B, device)

        # WF from two independent 9-frame raw means (no demodulation)
        wf_A = np.mean(raw9_A, axis=0).astype(np.float32)
        wf_B = np.mean(raw9_B, axis=0).astype(np.float32)

        frc_wf = compute_frc(wf_A, wf_B)
        frc_k3 = compute_frc(rec3_A, rec3_B)
        frc_k6 = compute_frc(rec6_A, rec6_B)
        frc_k9 = compute_frc(rec9_A, rec9_B)

        def resolve_nm(d: Dict[str, np.ndarray], which: str):
            thr = d["thr_halfbit"] if which == "half" else d["thr_1over7"]
            f_cut = estimate_cutoff(d["f_cyc_per_pix"], d["frc_smooth"], thr)
            return cutoff_to_nm(f_cut)

        row = {
            "sample": stem,
            "WF_halfbit_nm": resolve_nm(frc_wf, "half"),
            "WF_1over7_nm": resolve_nm(frc_wf, "17"),
            "K3_halfbit_nm": resolve_nm(frc_k3, "half"),
            "K3_1over7_nm": resolve_nm(frc_k3, "17"),
            "K6_halfbit_nm": resolve_nm(frc_k6, "half"),
            "K6_1over7_nm": resolve_nm(frc_k6, "17"),
            "K9_halfbit_nm": resolve_nm(frc_k9, "half"),
            "K9_1over7_nm": resolve_nm(frc_k9, "17"),
        }
        results.append(row)
        ws.append(list(row.values()))

        if stem == CFG.REP_SAMPLE_BASENAME:
            rep_curves = {
                "WF": frc_wf,
                "RL-SIM-3": frc_k3,
                "RL-SIM-6": frc_k6,
                "RL-SIM-9": frc_k9,
            }
            rep_nm = {
                "WF": row["WF_halfbit_nm"],
                "RL-SIM-3": row["K3_halfbit_nm"],
                "RL-SIM-6": row["K6_halfbit_nm"],
                "RL-SIM-9": row["K9_halfbit_nm"],
            }

    autosize_worksheet(ws)
    xlsx_path = out_dir / "paired_frc_results.xlsx"
    wb.save(xlsx_path)

    # CSV export
    import csv
    csv_path = out_dir / "paired_frc_per_sample.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(results[0].keys()))
        writer.writeheader()
        writer.writerows(results)

    # Main-text Fig. 1(g)
    draw_fig1g_boxplot(results, out_dir / "fig1g_frc_resolution_halfbit_nm.png")

    # Supplementary representative overlay
    if rep_curves:
        draw_rep_overlay(rep_curves, rep_nm, out_dir / "supp_fig_s1_representative_paired_frc.png")
        # also export representative source data
        rep_csv = out_dir / "paired_frc_source_curves.csv"
        import csv as _csv
        names = ["WF", "RL-SIM-3", "RL-SIM-6", "RL-SIM-9"]
        ref = rep_curves[names[0]]
        with open(rep_csv, "w", newline="", encoding="utf-8") as f:
            w = _csv.writer(f)
            header = ["f_norm"]
            for n in names:
                header += [f"{n}_frc_smooth", f"{n}_thr_halfbit", f"{n}_thr_1over7"]
            w.writerow(header)
            for i in range(len(ref["f_norm"])):
                row = [float(ref["f_norm"][i])]
                for n in names:
                    row += [
                        float(rep_curves[n]["frc_smooth"][i]) if np.isfinite(rep_curves[n]["frc_smooth"][i]) else "",
                        float(rep_curves[n]["thr_halfbit"][i]) if np.isfinite(rep_curves[n]["thr_halfbit"][i]) else "",
                        float(rep_curves[n]["thr_1over7"][i]) if np.isfinite(rep_curves[n]["thr_1over7"][i]) else "",
                    ]
                w.writerow(row)

    print("Saved:")
    print("  ", xlsx_path)
    print("  ", csv_path)
    print("  ", out_dir / "fig1g_frc_resolution_halfbit_nm.png")
    if rep_curves:
        print("  ", out_dir / "supp_fig_s1_representative_paired_frc.png")
        print("  ", out_dir / "paired_frc_source_curves.csv")


if __name__ == "__main__":
    main()
