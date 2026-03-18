# -*- coding: utf-8 -*-
"""
baseline_paired_frc_comparison.py

当前脚本实际执行的是：
GT-referenced FRC-style comparison on the held-out subset

新增功能
--------
本版会在 Excel 中额外生成：
1) fig_s3_curve_wide
   —— 直接用于重绘整张 supp_fig_s3_frc_vs_gt_subset 曲线图

2) fig_s3_boxplot_wide
   —— 直接用于重绘分辨率箱线图

3) curve_manifest
   —— 记录每个“单图-单算法”曲线 sheet 的名称、图像名、方法名和关键统计量

4) 4 × N 张单独曲线 sheet
   —— 每个算法、每张图各一张曲线表，格式接近你截图中的那种 Origin/Excel 曲线数据表
"""

from __future__ import annotations

import os
import sys
import math
import json
import random
import importlib.util
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import tifffile as tiff

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

import torch
import torch.nn as nn

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# =============================================================================
# Paths
# =============================================================================
ROOT = Path(r"E:\code\RL-SIM1")
GT_DIR = Path(r"C:\Users\1\OneDrive\Desktop\ceshiji")
CRSIM_ROOT_CANDS = [Path(r"E:\code\crsim6"), Path(r"E:\crsim6")]
RCAN_ROOT_CANDS = [Path(r"E:\code\RCAN-6"), Path(r"E:\RCAN-6")]

if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import infer_single_k369_and_frc as rlsim
from model_recon_scunet import SCUNetRecon


# =============================================================================
# Config
# =============================================================================
class CFG:
    DEVICE = "cuda" if torch.cuda.is_available() else "cpu"
    CHANNELS_LAST = True
    SEED = 2026

    # 若想固定 subset，可在这里写死文件名列表；None 表示读取 GT_DIR 下全部 tif
    GT_LIST = None
    GT_DIR = GT_DIR

    OUT_DIR = ROOT / "frc_vs_gt_baseline_subset"

    # 与当前 shared synthetic protocol 保持一致
    PATCH_SIZE = 256
    BASE_CYCLES = rlsim.CFG.BASE_CYCLES
    OTF_SIGMA = rlsim.CFG.OTF_SIGMA
    PIXEL_SIZE_NM = rlsim.CFG.PIXEL_SIZE_NM

    PHOTONS_PEAK = 500.0
    READ_STD_NORM = 0.003

    RL_CKPT = rlsim.CFG.CKPT_K6

    APOD_PX = 20
    REMOVE_MEAN = True
    MIN_RING_SAMPLES = 64
    SMOOTH_WIN = 9
    INTERP_GRID_N = 512
    DPI = 300

    SAVE_RECONS = True
    SAVE_RAW = False

    WF_CLIP_TO_01 = True
    RL_CLIP_TO_01 = True
    CR_CLIP_TO_01 = True
    RCAN_CLIP_TO_01 = True


METHOD_ORDER = ["WF", "RL-SIM-6", "CR-SIM-6", "RCAN-6"]
METHOD_COLORS = {
    "WF": "#7f7f7f",
    "RL-SIM-6": "#d62728",
    "CR-SIM-6": "#2ca02c",
    "RCAN-6": "#1f77b4",
}


# =============================================================================
# Helper utilities
# =============================================================================
def set_seed(seed: int):
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(seed)

def ensure_dir(path: Path | str) -> Path:
    p = Path(path)
    p.mkdir(parents=True, exist_ok=True)
    return p

def autosize_worksheet(ws):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 2, 42)

def list_tifs(folder: Path) -> List[Path]:
    exts = {".tif", ".tiff", ".TIF", ".TIFF"}
    if not folder.exists():
        raise FileNotFoundError(f"GT_DIR not found: {folder}")
    files = sorted([p for p in folder.iterdir() if p.suffix in exts])
    if not files:
        raise RuntimeError(f"No tif files found in: {folder}")
    return files

def get_gt_paths() -> List[Path]:
    if CFG.GT_LIST is not None:
        return [CFG.GT_DIR / x for x in CFG.GT_LIST]
    return list_tifs(CFG.GT_DIR)

def import_from_file(module_name: str, file_path: Path):
    spec = importlib.util.spec_from_file_location(module_name, str(file_path))
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module

def parse_ckpt_from_py(py_path: Path) -> Optional[Path]:
    if not py_path.exists():
        return None
    text = py_path.read_text(encoding="utf-8", errors="ignore")
    m = re.search(r'CKPT\s*=\s*pathlib\.Path\(r?"([^"]+)"\)', text)
    if not m:
        m = re.search(r'CKPT\s*=\s*r?"([^"]+)"', text)
    if m:
        return Path(m.group(1))
    return None

def find_existing_file(candidates: List[Path]) -> Path:
    for p in candidates:
        if p.exists():
            return p
    raise FileNotFoundError("No file found among candidates:\n" + "\n".join([str(x) for x in candidates]))

def save_tif(path: Path, arr: np.ndarray):
    path.parent.mkdir(parents=True, exist_ok=True)
    tiff.imwrite(str(path), arr.astype(np.float32))

def classify_structure(stem: str) -> str:
    s = stem.lower()
    if "microtubules" in s:
        return "MT"
    if "ccp" in s:
        return "CCP"
    if "er_" in s or "gtsim_level_06" in s:
        return "ER"
    return "Unknown"

def safe_method_key(method: str) -> str:
    return method.replace("-", "_").replace(" ", "_")

def compact_image_code(stem: str) -> str:
    """
    将长文件名压缩成适合 sheet 名的短码
    例如：
      CCPs_Cell_001_SIM_gt -> CCP001
      ER_Cell_068_GTSIM_level_06 -> ER068
      microtubules_Cell_055_SIM_gt -> MT055
    """
    m = re.search(r"CCPs?_Cell_(\d+)", stem, flags=re.IGNORECASE)
    if m:
        return f"CCP{int(m.group(1)):03d}"

    m = re.search(r"ER_Cell_(\d+)", stem, flags=re.IGNORECASE)
    if m:
        return f"ER{int(m.group(1)):03d}"

    m = re.search(r"microtubules_Cell_(\d+)", stem, flags=re.IGNORECASE)
    if m:
        return f"MT{int(m.group(1)):03d}"

    s = re.sub(r"[^A-Za-z0-9]+", "", stem)
    return s[:10] if len(s) > 10 else s

def short_method_tag(method: str) -> str:
    if method == "WF":
        return "WF"
    if method == "RL-SIM-6":
        return "RL6"
    if method == "CR-SIM-6":
        return "CR6"
    if method == "RCAN-6":
        return "RC6"
    return safe_method_key(method)[:4]

def make_curve_sheet_name(sample_idx: int, method: str, stem: str) -> str:
    """
    Excel sheet 名最长 31 字符
    """
    name = f"S{sample_idx:02d}_{short_method_tag(method)}_{compact_image_code(stem)}"
    return name[:31]


# =============================================================================
# Shared synthetic raw generator
# =============================================================================
@torch.no_grad()
def simulate_clean_raw6_phase_major(gt_path: Path, device: torch.device) -> Tuple[np.ndarray, np.ndarray]:
    arr = rlsim.imread(str(gt_path))
    x01 = rlsim.norm01(arr)
    ten = rlsim.to_tensor_1chw(x01)
    ten = rlsim.center_crop(ten, CFG.PATCH_SIZE)
    gt = ten.unsqueeze(0).to(device)
    if CFG.CHANNELS_LAST:
        gt = gt.contiguous(memory_format=torch.channels_last)
    forward = rlsim.SIMForwardAngles([0.0, 60.0], CFG.BASE_CYCLES, CFG.OTF_SIGMA).to(device).eval()
    raw = forward(gt)
    return raw[0].detach().cpu().numpy().astype(np.float32), gt[0, 0].detach().cpu().numpy().astype(np.float32)

def add_poisson_gaussian_noise(raw_clean: np.ndarray, photons_peak: float, read_std_norm: float, seed: int) -> np.ndarray:
    rng = np.random.default_rng(seed)
    x = np.clip(raw_clean.astype(np.float64, copy=False), 0.0, 1.0)
    lam = photons_peak * x
    shot = rng.poisson(lam).astype(np.float64) / float(photons_peak)
    noisy = shot + rng.normal(0.0, read_std_norm, size=x.shape)
    return np.clip(noisy, 0.0, 1.0).astype(np.float32)


# =============================================================================
# Channel reorder
# =============================================================================
def phase_major_to_angle_major_6(raw_pm: np.ndarray) -> np.ndarray:
    """
    phase-major:
      [a0/p0, a60/p0, a0/p120, a60/p120, a0/p240, a60/p240]
    angle-major:
      [a0/p0, a0/p120, a0/p240, a60/p0, a60/p120, a60/p240]
    """
    idx = [0, 2, 4, 1, 3, 5]
    return raw_pm[idx].copy()


# =============================================================================
# Model loading
# =============================================================================
def discover_crsim_root() -> Path:
    for p in CRSIM_ROOT_CANDS:
        if (p / "eval_single_gt.py").exists() and (p / "models" / "crsim6_model_big.py").exists():
            return p
    raise FileNotFoundError("Could not locate CR-SIM root.")

def discover_rcan_root() -> Path:
    for p in RCAN_ROOT_CANDS:
        if (p / "models_rcan6f.py").exists():
            return p
    raise FileNotFoundError("Could not locate RCAN-6 root.")

def load_rlsim6(device: torch.device) -> nn.Module:
    net = rlsim.load_scunet(str(CFG.RL_CKPT), 6, device)
    net.eval()
    return net

def load_crsim6(device: torch.device) -> Tuple[nn.Module, Dict[str, str]]:
    root = discover_crsim_root()
    ckpt = parse_ckpt_from_py(root / "eval_single_gt.py")
    if ckpt is None:
        ckpt = parse_ckpt_from_py(root / "infer.py")
    if ckpt is None or not ckpt.exists():
        raise FileNotFoundError("Could not discover CR-SIM checkpoint.")

    mod = import_from_file("crsim6_model_big_local", root / "models" / "crsim6_model_big.py")
    NetCls = None
    for name in ["CRSIMNetBig6", "CRSIMNet"]:
        if hasattr(mod, name):
            NetCls = getattr(mod, name)
            break
    if NetCls is None:
        raise AttributeError("Neither CRSIMNetBig6 nor CRSIMNet found in crsim6_model_big.py")

    net = NetCls().to(device)
    state = torch.load(str(ckpt), map_location=device)
    net.load_state_dict(state, strict=False)
    net.eval()

    info = {
        "CRSIM_ROOT": str(root),
        "CRSIM_CKPT": str(ckpt),
        "CRSIM_CLASS": NetCls.__name__,
    }
    return net, info

def load_rcan6(device: torch.device) -> Tuple[nn.Module, Dict[str, str]]:
    root = discover_rcan_root()
    ckpt = find_existing_file([
        root / "checkpoints" / "best.pth",
        root / "checkpoints" / "last.pth",
    ])

    mod = import_from_file("models_rcan6f_local", root / "models_rcan6f.py")
    NetCls = None
    for name in ["RCAN6F", "RCAN", "Model"]:
        if hasattr(mod, name):
            NetCls = getattr(mod, name)
            break
    if NetCls is None:
        raise AttributeError("Could not find RCAN model class in models_rcan6f.py")

    net = NetCls().to(device)
    state = torch.load(str(ckpt), map_location=device)
    if isinstance(state, dict) and "model" in state:
        state = state["model"]
    net.load_state_dict(state, strict=False)
    net.eval()

    info = {
        "RCAN_ROOT": str(root),
        "RCAN_CKPT": str(ckpt),
        "RCAN_CLASS": NetCls.__name__,
    }
    return net, info


# =============================================================================
# Reconstruction wrappers
# =============================================================================
@torch.no_grad()
def run_wf_from_raw(raw_pm: np.ndarray) -> np.ndarray:
    wf = np.mean(raw_pm, axis=0).astype(np.float32)
    if CFG.WF_CLIP_TO_01:
        wf = np.clip(wf, 0.0, 1.0)
    return wf

@torch.no_grad()
def run_rlsim6_from_raw(net: nn.Module, raw_pm: np.ndarray, device: torch.device) -> np.ndarray:
    inp = torch.from_numpy(raw_pm[None]).to(device)
    if CFG.CHANNELS_LAST:
        inp = inp.contiguous(memory_format=torch.channels_last)
    rec = torch.sigmoid(net(inp)).clamp(0.0, 1.0)
    out = rec[0, 0].detach().cpu().numpy().astype(np.float32)
    if CFG.RL_CLIP_TO_01:
        out = np.clip(out, 0.0, 1.0)
    return out

@torch.no_grad()
def run_crsim6_from_raw(net: nn.Module, raw_pm: np.ndarray, device: torch.device, target_hw: Tuple[int, int]) -> np.ndarray:
    raw_am = phase_major_to_angle_major_6(raw_pm)
    inp = torch.from_numpy(raw_am[None]).to(device)
    sr = net(inp).squeeze().detach().cpu().numpy().astype(np.float32)

    if sr.shape != target_hw:
        try:
            import cv2
            sr = cv2.resize(sr, target_hw[::-1], interpolation=cv2.INTER_CUBIC)
        except Exception:
            Ht, Wt = target_hw
            H, W = sr.shape
            out = np.zeros((Ht, Wt), dtype=np.float32)
            h = min(H, Ht)
            w = min(W, Wt)
            ys = (H - h) // 2
            xs = (W - w) // 2
            yt = (Ht - h) // 2
            xt = (Wt - w) // 2
            out[yt:yt+h, xt:xt+w] = sr[ys:ys+h, xs:xs+w]
            sr = out

    if CFG.CR_CLIP_TO_01:
        sr = np.clip(sr, 0.0, 1.0)
    return sr

@torch.no_grad()
def run_rcan6_from_raw(net: nn.Module, raw_pm: np.ndarray, device: torch.device, target_hw: Tuple[int, int]) -> np.ndarray:
    inp = torch.from_numpy(raw_pm[None]).to(device)
    sr = net(inp).squeeze().detach().cpu().numpy().astype(np.float32)

    if sr.shape != target_hw:
        try:
            import cv2
            sr = cv2.resize(sr, target_hw[::-1], interpolation=cv2.INTER_CUBIC)
        except Exception:
            Ht, Wt = target_hw
            H, W = sr.shape
            out = np.zeros((Ht, Wt), dtype=np.float32)
            h = min(H, Ht)
            w = min(W, Wt)
            ys = (H - h) // 2
            xs = (W - w) // 2
            yt = (Ht - h) // 2
            xt = (Wt - w) // 2
            out[yt:yt+h, xt:xt+w] = sr[ys:ys+h, xs:xs+w]
            sr = out

    if CFG.RCAN_CLIP_TO_01:
        sr = np.clip(sr, 0.0, 1.0)
    return sr


# =============================================================================
# FRC vs GT
# =============================================================================
def cosine_apodization_window(h: int, w: int, apod_px: int) -> np.ndarray:
    if apod_px <= 0:
        return np.ones((h, w), dtype=np.float64)
    ap = int(apod_px)
    ap = min(ap, h // 2 - 1, w // 2 - 1)
    if ap <= 0:
        return np.ones((h, w), dtype=np.float64)
    t = 0.5 * (1.0 - np.cos(np.linspace(0.0, math.pi, ap, dtype=np.float64)))
    wy = np.ones(h, dtype=np.float64)
    wx = np.ones(w, dtype=np.float64)
    wy[:ap] = t
    wy[-ap:] = t[::-1]
    wx[:ap] = t
    wx[-ap:] = t[::-1]
    return wy[:, None] * wx[None, :]

def smooth_1d_nan(x: np.ndarray, win: int) -> np.ndarray:
    if win <= 1:
        return x.astype(np.float64, copy=True)
    y = np.full_like(x, np.nan, dtype=np.float64)
    n = x.size
    half = win // 2
    for i in range(n):
        j0 = max(0, i - half)
        j1 = min(n, i + half + 1)
        seg = x[j0:j1]
        seg = seg[np.isfinite(seg)]
        if seg.size > 0:
            y[i] = float(np.mean(seg))
    return y

def half_bit_threshold(Ni: np.ndarray) -> np.ndarray:
    Ni = Ni.astype(np.float64)
    T = np.full_like(Ni, np.nan, dtype=np.float64)
    m = Ni > 0
    s = np.sqrt(Ni[m])
    T[m] = (0.2071 + 1.9102 / s) / (1.2701 + 0.9102 / s)
    return T

def compute_frc(img_ref: np.ndarray, img_tst: np.ndarray, apod_px: int, remove_mean: bool, min_ring_samples: int, smooth_win: int):
    if img_ref.shape != img_tst.shape:
        raise ValueError(f"Shape mismatch: ref={img_ref.shape}, tst={img_tst.shape}")

    ref = np.nan_to_num(img_ref.astype(np.float64), nan=0.0, posinf=0.0, neginf=0.0)
    tst = np.nan_to_num(img_tst.astype(np.float64), nan=0.0, posinf=0.0, neginf=0.0)

    if remove_mean:
        ref = ref - float(np.mean(ref))
        tst = tst - float(np.mean(tst))

    H, W = ref.shape
    win = cosine_apodization_window(H, W, apod_px)
    refw = ref * win
    tstw = tst * win

    F1 = np.fft.fftshift(np.fft.fft2(refw))
    F2 = np.fft.fftshift(np.fft.fft2(tstw))

    fy = (np.arange(H) - (H // 2)) / float(H)
    fx = (np.arange(W) - (W // 2)) / float(W)
    FX, FY = np.meshgrid(fx, fy)
    FR = np.sqrt(FX * FX + FY * FY)

    f_max = 0.5
    m = FR <= f_max
    FRv = FR[m].ravel()
    F1v = F1[m].ravel()
    F2v = F2[m].ravel()

    Nmin = min(H, W)
    n_bins = max(int(np.floor(f_max * Nmin)), 8)

    bin_idx = np.floor(FRv / f_max * n_bins).astype(np.int64)
    bin_idx = np.clip(bin_idx, 0, n_bins - 1)

    Ni = np.bincount(bin_idx, minlength=n_bins).astype(np.int64)
    cross = np.real(F1v * np.conj(F2v))
    p1 = np.abs(F1v) ** 2
    p2 = np.abs(F2v) ** 2

    num = np.bincount(bin_idx, weights=cross, minlength=n_bins).astype(np.float64)
    den1 = np.bincount(bin_idx, weights=p1, minlength=n_bins).astype(np.float64)
    den2 = np.bincount(bin_idx, weights=p2, minlength=n_bins).astype(np.float64)

    eps = 1e-12
    frc = num / np.sqrt(den1 * den2 + eps)
    frc = frc.astype(np.float64)
    frc[Ni < int(min_ring_samples)] = np.nan

    f_cyc_per_pix = (np.arange(n_bins, dtype=np.float64) + 0.5) * f_max / float(n_bins)
    f_norm = f_cyc_per_pix / f_max
    thr_half = half_bit_threshold(Ni.astype(np.float64))
    thr_17 = np.full(n_bins, 1.0 / 7.0, dtype=np.float64)
    frc_s = smooth_1d_nan(frc, smooth_win)

    return {
        "f_norm": f_norm,
        "f_cyc_per_pix": f_cyc_per_pix,
        "frc_raw": frc,
        "frc_smooth": frc_s,
        "thr_halfbit": thr_half,
        "thr_1over7": thr_17,
        "Ni": Ni.astype(np.float64),
    }

def estimate_cutoff_first_crossing(f: np.ndarray, frc: np.ndarray, thr: np.ndarray) -> Optional[float]:
    valid = np.isfinite(f) & np.isfinite(frc) & np.isfinite(thr)
    if np.sum(valid) < 2:
        return None
    fv = f[valid]
    gv = frc[valid] - thr[valid]

    if gv[0] < 0:
        return None

    for i in range(len(fv) - 1):
        g0 = gv[i]
        g1 = gv[i + 1]
        if (g0 >= 0.0) and (g1 < 0.0):
            x0, x1 = fv[i], fv[i + 1]
            if g0 == g1:
                return float(x0)
            t = g0 / (g0 - g1)
            t = min(max(t, 0.0), 1.0)
            return float(x0 + t * (x1 - x0))

    if np.all(gv >= 0.0):
        return float(fv[-1])

    return None

def freq_to_resolution_nm(f_cut: Optional[float]) -> Optional[float]:
    if f_cut is None or (not np.isfinite(f_cut)) or f_cut <= 0:
        return None
    return float((1.0 / f_cut) * CFG.PIXEL_SIZE_NM)

def auc_curve(f: np.ndarray, y: np.ndarray) -> float:
    valid = np.isfinite(f) & np.isfinite(y)
    if np.sum(valid) < 2:
        return float("nan")
    return float(np.trapz(y[valid], f[valid]))


# =============================================================================
# Main
# =============================================================================
def main():
    set_seed(CFG.SEED)
    device = torch.device(CFG.DEVICE)
    out_dir = ensure_dir(CFG.OUT_DIR)
    recon_dir = ensure_dir(out_dir / "per_sample_recons")

    gt_paths = get_gt_paths()
    print(f"[Info] device = {device}")
    print(f"[Info] n_images = {len(gt_paths)}")
    print(f"[Info] GT_DIR = {CFG.GT_DIR}")
    print(f"[Info] RL_CKPT = {CFG.RL_CKPT}")

    # load models once
    print("[Info] loading RL-SIM-6 ...")
    rl_net = load_rlsim6(device)

    print("[Info] loading CR-SIM-6 ...")
    cr_net, cr_info = load_crsim6(device)

    print("[Info] loading RCAN-6 ...")
    rc_net, rc_info = load_rcan6(device)

    discovered = {"RL_CKPT_K6": str(CFG.RL_CKPT), "GT_DIR": str(CFG.GT_DIR)}
    discovered.update(cr_info)
    discovered.update(rc_info)

    interp_grid = np.linspace(0.0, 1.0, CFG.INTERP_GRID_N, dtype=np.float64)
    curve_bank: Dict[str, List[np.ndarray]] = {m: [] for m in METHOD_ORDER}
    per_image_rows = []
    curve_detail_records = []

    for i, gt_path in enumerate(gt_paths, start=1):
        stem = gt_path.stem
        print(f"\n[Sample {i}/{len(gt_paths)}] {stem}")
        sample_dir = ensure_dir(recon_dir / stem)

        raw_clean_pm, gt_crop = simulate_clean_raw6_phase_major(gt_path, device)
        Ht, Wt = gt_crop.shape

        raw_noisy_pm = add_poisson_gaussian_noise(
            raw_clean_pm, CFG.PHOTONS_PEAK, CFG.READ_STD_NORM, CFG.SEED + 1000 + i
        )

        if CFG.SAVE_RAW:
            save_tif(sample_dir / "raw_noisy_phase_major.tif", raw_noisy_pm)

        recs = {}
        recs["WF"] = run_wf_from_raw(raw_noisy_pm)
        recs["RL-SIM-6"] = run_rlsim6_from_raw(rl_net, raw_noisy_pm, device)
        recs["CR-SIM-6"] = run_crsim6_from_raw(cr_net, raw_noisy_pm, device, target_hw=(Ht, Wt))
        recs["RCAN-6"] = run_rcan6_from_raw(rc_net, raw_noisy_pm, device, target_hw=(Ht, Wt))

        if CFG.SAVE_RECONS:
            save_tif(sample_dir / "gt_crop.tif", gt_crop)
            for method in METHOD_ORDER:
                safe = safe_method_key(method)
                save_tif(sample_dir / f"{safe}.tif", recs[method])

        for method in METHOD_ORDER:
            out = compute_frc(
                img_ref=gt_crop,
                img_tst=recs[method],
                apod_px=CFG.APOD_PX,
                remove_mean=CFG.REMOVE_MEAN,
                min_ring_samples=CFG.MIN_RING_SAMPLES,
                smooth_win=CFG.SMOOTH_WIN,
            )

            f_norm = out["f_norm"]
            f = out["f_cyc_per_pix"]
            frc_raw = out["frc_raw"]
            frc_s = out["frc_smooth"]
            thr_half = out["thr_halfbit"]
            thr_17 = out["thr_1over7"]
            Ni = out["Ni"]

            cut_half = estimate_cutoff_first_crossing(f, frc_s, thr_half)
            cut_17 = estimate_cutoff_first_crossing(f, frc_s, thr_17)

            res_half_nm = freq_to_resolution_nm(cut_half)
            res_17_nm = freq_to_resolution_nm(cut_17)
            auc_val = auc_curve(f_norm, frc_s)

            note = ""
            valid = np.isfinite(f) & np.isfinite(frc_s) & np.isfinite(thr_17)
            if np.sum(valid) >= 2:
                gv = frc_s[valid] - thr_17[valid]
                if np.all(gv >= 0.0):
                    note = "curve stayed above 1/7 threshold to final valid bin"

            per_image_rows.append({
                "image": stem,
                "structure": classify_structure(stem),
                "method": method,
                "cutoff_halfbit_cyc_per_px": cut_half,
                "resolution_halfbit_nm": res_half_nm,
                "cutoff_1over7_cyc_per_px": cut_17,
                "resolution_1over7_nm": res_17_nm,
                "auc": auc_val,
                "note": note,
            })

            curve_detail_records.append({
                "sample_idx": i,
                "image": stem,
                "structure": classify_structure(stem),
                "method": method,
                "sheet_name": make_curve_sheet_name(i, method, stem),
                "cutoff_halfbit_cyc_per_px": cut_half,
                "resolution_halfbit_nm": res_half_nm,
                "cutoff_1over7_cyc_per_px": cut_17,
                "resolution_1over7_nm": res_17_nm,
                "auc": auc_val,
                "note": note,
                "f_norm": f_norm.copy(),
                "f_cyc_per_pix": f.copy(),
                "frc_raw": frc_raw.copy(),
                "frc_smooth": frc_s.copy(),
                "thr_halfbit": thr_half.copy(),
                "thr_1over7": thr_17.copy(),
                "Ni": Ni.copy(),
            })

            valid_curve = np.isfinite(f_norm) & np.isfinite(frc_s)
            if np.sum(valid_curve) >= 2:
                x = f_norm[valid_curve]
                y = frc_s[valid_curve]
                curve_interp = np.interp(interp_grid, x, y, left=y[0], right=y[-1])
            elif np.sum(valid_curve) == 1:
                y0 = float(frc_s[valid_curve][0])
                curve_interp = np.full_like(interp_grid, y0, dtype=np.float64)
            else:
                curve_interp = np.full_like(interp_grid, np.nan, dtype=np.float64)

            curve_bank[method].append(curve_interp)

            print(
                f"  [{method}] "
                f"half-bit={res_half_nm if res_half_nm is not None else 'NA'} nm | "
                f"1/7={res_17_nm if res_17_nm is not None else 'NA'} nm | "
                f"AUC={auc_val:.4f}"
                + (f" | note={note}" if note else "")
            )

    # aggregate
    mean_curve_rows = []
    summary_rows = []
    curve_wide_rows = []
    curve_stats = {}

    plt.figure(figsize=(8.0, 5.8))
    for method in METHOD_ORDER:
        mat = np.stack(curve_bank[method], axis=0)
        mean_curve = np.nanmean(mat, axis=0)
        std_curve = np.nanstd(mat, axis=0)
        curve_stats[method] = {
            "mean": mean_curve,
            "std": std_curve,
            "lower": mean_curve - std_curve,
            "upper": mean_curve + std_curve,
        }

        for f0, mu, sd in zip(interp_grid, mean_curve, std_curve):
            mean_curve_rows.append({
                "method": method,
                "freq_norm": float(f0),
                "frc_mean": float(mu),
                "frc_std": float(sd),
            })

        vals_half = [
            r["resolution_halfbit_nm"] for r in per_image_rows
            if (r["method"] == method) and (r["resolution_halfbit_nm"] is not None) and np.isfinite(r["resolution_halfbit_nm"])
        ]
        vals_17 = [
            r["resolution_1over7_nm"] for r in per_image_rows
            if (r["method"] == method) and (r["resolution_1over7_nm"] is not None) and np.isfinite(r["resolution_1over7_nm"])
        ]
        aucs = [
            r["auc"] for r in per_image_rows
            if (r["method"] == method) and np.isfinite(r["auc"])
        ]

        q1 = float(np.percentile(vals_17, 25)) if len(vals_17) else None
        q3 = float(np.percentile(vals_17, 75)) if len(vals_17) else None

        summary_rows.append({
            "method": method,
            "n_images": len([r for r in per_image_rows if r["method"] == method]),
            "mean_resolution_halfbit_nm": float(np.mean(vals_half)) if len(vals_half) else None,
            "median_resolution_halfbit_nm": float(np.median(vals_half)) if len(vals_half) else None,
            "mean_resolution_1over7_nm": float(np.mean(vals_17)) if len(vals_17) else None,
            "median_resolution_1over7_nm": float(np.median(vals_17)) if len(vals_17) else None,
            "iqr_q1_resolution_1over7_nm": q1,
            "iqr_q3_resolution_1over7_nm": q3,
            "mean_auc": float(np.mean(aucs)) if len(aucs) else None,
        })

        plt.plot(interp_grid, mean_curve, linewidth=2.2, color=METHOD_COLORS[method], label=method)
        plt.fill_between(interp_grid, mean_curve - std_curve, mean_curve + std_curve,
                         color=METHOD_COLORS[method], alpha=0.18, linewidth=0)

    # 直接重绘整张图的宽表
    for idx, f0 in enumerate(interp_grid):
        row = {
            "freq_norm": float(f0),
            "threshold_1over7": float(1.0 / 7.0),
        }
        for method in METHOD_ORDER:
            key = safe_method_key(method)
            row[f"{key}_mean"] = float(curve_stats[method]["mean"][idx])
            row[f"{key}_std"] = float(curve_stats[method]["std"][idx])
            row[f"{key}_lower"] = float(curve_stats[method]["lower"][idx])
            row[f"{key}_upper"] = float(curve_stats[method]["upper"][idx])
        curve_wide_rows.append(row)

    plt.axhline(1.0 / 7.0, linestyle="--", linewidth=1.4, color="black", label="1/7 threshold")
    plt.xlabel("Normalized spatial frequency (Nyquist=1)")
    plt.ylabel("GT-referenced FRC-style correlation")
    plt.title("GT-referenced FRC-style comparison on the held-out subset")
    plt.xlim(0.0, 1.0)
    plt.ylim(-0.02, 1.02)
    plt.grid(True, linewidth=0.4, alpha=0.45)
    plt.legend(frameon=False, fontsize=9)
    plt.tight_layout()
    plt.savefig(str(out_dir / "supp_fig_s3_frc_vs_gt_subset.png"), dpi=CFG.DPI)
    plt.savefig(str(out_dir / "supp_fig_s3_frc_vs_gt_subset.pdf"))
    plt.close()

    # boxplot
    plt.figure(figsize=(8.0, 5.8))
    data_box = []
    for method in METHOD_ORDER:
        vals = [
            r["resolution_1over7_nm"] for r in per_image_rows
            if (r["method"] == method) and (r["resolution_1over7_nm"] is not None) and np.isfinite(r["resolution_1over7_nm"])
        ]
        data_box.append(vals)

    bp = plt.boxplot(data_box, labels=METHOD_ORDER, showfliers=False, patch_artist=True)
    for patch, method in zip(bp["boxes"], METHOD_ORDER):
        patch.set_facecolor(METHOD_COLORS[method])
        patch.set_alpha(0.28)

    for i, method in enumerate(METHOD_ORDER, start=1):
        vals = [
            r["resolution_1over7_nm"] for r in per_image_rows
            if (r["method"] == method) and (r["resolution_1over7_nm"] is not None) and np.isfinite(r["resolution_1over7_nm"])
        ]
        if len(vals):
            x = np.full(len(vals), i, dtype=np.float32)
            jitter = np.linspace(-0.08, 0.08, len(vals), dtype=np.float32) if len(vals) > 1 else np.array([0.0], dtype=np.float32)
            plt.scatter(x + jitter, vals, s=28, alpha=0.9, color=METHOD_COLORS[method], edgecolors="none")

    plt.ylabel("GT-referenced FRC-derived resolution (nm, 1/7)")
    plt.title("Per-image GT-referenced FRC-derived resolution on the held-out subset")
    plt.grid(True, axis="y", linewidth=0.4, alpha=0.45)
    plt.tight_layout()
    plt.savefig(str(out_dir / "supp_fig_s3_resolution_boxplot_vs_gt_subset.png"), dpi=CFG.DPI)
    plt.savefig(str(out_dir / "supp_fig_s3_resolution_boxplot_vs_gt_subset.pdf"))
    plt.close()

    # discovered paths
    with open(out_dir / "discovered_paths.json", "w", encoding="utf-8") as f:
        json.dump(discovered, f, indent=2, ensure_ascii=False)

    # 直接重绘 boxplot 的宽表
    per_image_wide = {}
    for row in per_image_rows:
        img = row["image"]
        if img not in per_image_wide:
            per_image_wide[img] = {
                "image": img,
                "structure": row["structure"],
            }
        key = safe_method_key(row["method"])
        per_image_wide[img][f"{key}_resolution_1over7_nm"] = row["resolution_1over7_nm"]
        per_image_wide[img][f"{key}_resolution_halfbit_nm"] = row["resolution_halfbit_nm"]
        per_image_wide[img][f"{key}_auc"] = row["auc"]

    per_image_wide_rows = [per_image_wide[k] for k in sorted(per_image_wide.keys())]

    # workbook
    wb = Workbook()

    # 1) summary
    ws1 = wb.active
    ws1.title = "summary"
    ws1.append([
        "method",
        "n_images",
        "mean_resolution_halfbit_nm",
        "median_resolution_halfbit_nm",
        "mean_resolution_1over7_nm",
        "median_resolution_1over7_nm",
        "iqr_q1_resolution_1over7_nm",
        "iqr_q3_resolution_1over7_nm",
        "mean_auc",
    ])
    for row in summary_rows:
        ws1.append([
            row["method"],
            row["n_images"],
            row["mean_resolution_halfbit_nm"],
            row["median_resolution_halfbit_nm"],
            row["mean_resolution_1over7_nm"],
            row["median_resolution_1over7_nm"],
            row["iqr_q1_resolution_1over7_nm"],
            row["iqr_q3_resolution_1over7_nm"],
            row["mean_auc"],
        ])
    autosize_worksheet(ws1)

    # 2) per_image_long
    ws2 = wb.create_sheet("per_image_long")
    ws2.append([
        "image",
        "structure",
        "method",
        "cutoff_halfbit_cyc_per_px",
        "resolution_halfbit_nm",
        "cutoff_1over7_cyc_per_px",
        "resolution_1over7_nm",
        "auc",
        "note",
    ])
    for row in per_image_rows:
        ws2.append([
            row["image"],
            row["structure"],
            row["method"],
            row["cutoff_halfbit_cyc_per_px"],
            row["resolution_halfbit_nm"],
            row["cutoff_1over7_cyc_per_px"],
            row["resolution_1over7_nm"],
            row["auc"],
            row["note"],
        ])
    autosize_worksheet(ws2)

    # 3) mean_curves_long
    ws3 = wb.create_sheet("mean_curves_long")
    ws3.append(["method", "freq_norm", "frc_mean", "frc_std"])
    for row in mean_curve_rows:
        ws3.append([row["method"], row["freq_norm"], row["frc_mean"], row["frc_std"]])
    autosize_worksheet(ws3)

    # 4) fig_s3_curve_wide —— 直接用于重绘整张 supp_fig_s3_frc_vs_gt_subset
    ws4 = wb.create_sheet("fig_s3_curve_wide")
    curve_headers = ["freq_norm", "threshold_1over7"]
    for method in METHOD_ORDER:
        key = safe_method_key(method)
        curve_headers.extend([
            f"{key}_mean",
            f"{key}_std",
            f"{key}_lower",
            f"{key}_upper",
        ])
    ws4.append(curve_headers)
    for row in curve_wide_rows:
        ws4.append([row.get(h, None) for h in curve_headers])
    autosize_worksheet(ws4)

    # 5) fig_s3_boxplot_wide —— 直接用于重绘 boxplot
    ws5 = wb.create_sheet("fig_s3_boxplot_wide")
    box_headers = ["image", "structure"]
    for method in METHOD_ORDER:
        key = safe_method_key(method)
        box_headers.extend([
            f"{key}_resolution_1over7_nm",
            f"{key}_resolution_halfbit_nm",
            f"{key}_auc",
        ])
    ws5.append(box_headers)
    for row in per_image_wide_rows:
        ws5.append([row.get(h, None) for h in box_headers])
    autosize_worksheet(ws5)

    # 6) curve_manifest —— 记录 60 个单独曲线 sheet
    ws6 = wb.create_sheet("curve_manifest")
    ws6.append([
        "sheet_name",
        "sample_idx",
        "image",
        "structure",
        "method",
        "cutoff_halfbit_cyc_per_px",
        "resolution_halfbit_nm",
        "cutoff_1over7_cyc_per_px",
        "resolution_1over7_nm",
        "auc",
        "note",
    ])
    for rec in curve_detail_records:
        ws6.append([
            rec["sheet_name"],
            rec["sample_idx"],
            rec["image"],
            rec["structure"],
            rec["method"],
            rec["cutoff_halfbit_cyc_per_px"],
            rec["resolution_halfbit_nm"],
            rec["cutoff_1over7_cyc_per_px"],
            rec["resolution_1over7_nm"],
            rec["auc"],
            rec["note"],
        ])
    autosize_worksheet(ws6)

    # 7) 为每个“图像×算法”生成单独曲线 sheet（4 × N）
    for rec in curve_detail_records:
        ws = wb.create_sheet(rec["sheet_name"])

        # 顶部元数据
        ws["A1"] = "method"
        ws["B1"] = rec["method"]

        ws["A2"] = "image"
        ws["B2"] = rec["image"]

        ws["A3"] = "structure"
        ws["B3"] = rec["structure"]

        ws["A4"] = "cutoff_halfbit_cyc_per_px"
        ws["B4"] = rec["cutoff_halfbit_cyc_per_px"]

        ws["A5"] = "resolution_halfbit_nm"
        ws["B5"] = rec["resolution_halfbit_nm"]

        ws["A6"] = "cutoff_1over7_cyc_per_px"
        ws["B6"] = rec["cutoff_1over7_cyc_per_px"]

        ws["A7"] = "resolution_1over7_nm"
        ws["B7"] = rec["resolution_1over7_nm"]

        ws["A8"] = "auc"
        ws["B8"] = rec["auc"]

        ws["A9"] = "note"
        ws["B9"] = rec["note"]

        # 数据头
        start_row = 11
        headers = [
            "f_norm(Nyq=1)",
            "f_cyc_per_pix",
            "FRC_raw",
            "FRC_smooth",
            "thr_halfbit",
            "thr_1over7",
            "Ni",
        ]
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=start_row, column=col_idx, value=h)

        # 数据
        n = len(rec["f_norm"])
        for irow in range(n):
            ws.cell(row=start_row + 1 + irow, column=1, value=float(rec["f_norm"][irow]) if np.isfinite(rec["f_norm"][irow]) else None)
            ws.cell(row=start_row + 1 + irow, column=2, value=float(rec["f_cyc_per_pix"][irow]) if np.isfinite(rec["f_cyc_per_pix"][irow]) else None)
            ws.cell(row=start_row + 1 + irow, column=3, value=float(rec["frc_raw"][irow]) if np.isfinite(rec["frc_raw"][irow]) else None)
            ws.cell(row=start_row + 1 + irow, column=4, value=float(rec["frc_smooth"][irow]) if np.isfinite(rec["frc_smooth"][irow]) else None)
            ws.cell(row=start_row + 1 + irow, column=5, value=float(rec["thr_halfbit"][irow]) if np.isfinite(rec["thr_halfbit"][irow]) else None)
            ws.cell(row=start_row + 1 + irow, column=6, value=float(rec["thr_1over7"][irow]) if np.isfinite(rec["thr_1over7"][irow]) else None)
            ws.cell(row=start_row + 1 + irow, column=7, value=float(rec["Ni"][irow]) if np.isfinite(rec["Ni"][irow]) else None)

        autosize_worksheet(ws)

    xlsx_path = out_dir / "supp_table_s5_frc_vs_gt_subset.xlsx"
    wb.save(str(xlsx_path))

    print("\n==================== DONE ====================")
    print(f"[save] {out_dir / 'supp_fig_s3_frc_vs_gt_subset.png'}")
    print(f"[save] {out_dir / 'supp_fig_s3_resolution_boxplot_vs_gt_subset.png'}")
    print(f"[save] {xlsx_path}")
    print(f"[save] {out_dir / 'discovered_paths.json'}")
    print("[Info] redraw whole-figure curve sheet: fig_s3_curve_wide")
    print("[Info] redraw whole-figure boxplot sheet: fig_s3_boxplot_wide")
    print("[Info] detailed per-image per-method curve sheets: curve_manifest + 4×N curve sheets")


if __name__ == "__main__":
    main()